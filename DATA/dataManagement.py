# -- coding: utf-8 --
"""Wed Aug  17 11:37:50 2020

@author: Macrov
"""


import pandas as pd
import os
from openpyxl import load_workbook
import pandas.io.formats.excel


def update_excel(data,file):
    try:
        """Configure pandas to load the file, remove the header cells style and save the file in the
        correct place"""                      
        book = load_workbook(os.path.join(os.path.dirname(os.getcwd()),"BD\{}".format(file)))
        writer = pd.ExcelWriter(os.path.join(os.path.dirname(os.getcwd()),"BD\{}".format(file)), engine='openpyxl') 
        writer.book = book
        
        pandas.io.formats.excel.header_style = None
        
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
        data.to_excel(writer, "Presupuesto", startrow = book.worksheets[0].max_row,header=None) 
        writer.save()
        status = "Success"
    except:
        status = "Failure"
    finally:
        return(status)

def read(file):
        """
        Reads the specified file
        and return a dataframe with the file info
        PARAMS:
            file: the name of the file(Database)
        RETURNS:
            Dataframe containing the file data
        """
        try:
            data = pd.read_excel(os.path.join(os.path.dirname(os.getcwd()),"BD\{}".format(file)), sheet_name=0)
            return data
        except:
            print("No such file")    
    
def export(data,file):
        """
        Reads the dataframe, export the new data
        and return if the export has beean a succes or a failure
        PARAMS:
            data: the dataframe
            file: File name of the exported data
            
        RETURNS:
            Succes or failure
        """
        state="Failure"
        if os.path.isfile(os.path.join(os.path.dirname(os.getcwd()),"BD\{}".format(file))):
            file="MacrovExported.xlsx"
            data = pd.DataFrame({'No. item':'21.6.0','Analisis de precios unitarios':"Prueba actualizacion","Unidad":"día","Importe de materiales":[10000],"Importe de mano de obra":[9999],"Importe de equipo":[132958],"Importe de auxiliares":[28761],"Importe de conceptos":[48361],"Precio unitario":[9999]})
            return update_excel(data,file)
        else: 
            try:
                dataframe = pd.DataFrame(data)
                dataframe.to_excel(os.path.join(os.path.dirname(os.getcwd()),"BD\{}".format(file)), sheet_name='Presupuesto')
                state="Success"
                return(state)
            except:
                return(state)    
            
    
def budget(data,productsIds, quantity):
        """
        Reads the dataframe and the id of the products wnated in that dataframe
        and return the total budget required for those products
        PARAMS:
            data: the dataframe
            productsId: the list of id's wnated for the budget
            quantity: the list of how much product is required
        RETURNS:
            The total budget
        """
        
        data=data.fillna(0)
        budg=0
        products=productsIds
        quantity=quantity
        
        dataFilter = data[data['No. Item'].isin(products)]
        prices=dataFilter['Precio unitario'].tolist()
        
        for i in range(len(quantity)):
            budg+=quantity[i]*prices[i]
        return budg 

file="MACRO.xlsx"
listProducts = ['1.1.1','2.2.2','2.2.3']
quantity=[1,2,3]
  
data=read(file)
bud=budget(data,listProducts,quantity)
fout="MacrovExported.xlsx"

print(bud)
print(export(data,fout))
